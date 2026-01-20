from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def set_cell_hyperlink(cell, url: str, label: str | None = None) -> None:
    if label is not None:
        cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def clear_hyperlink(cell) -> None:
    cell.hyperlink = None


def rewrite_overview(wb, project_name: str, deliverables_path: str) -> None:
    if "Overview" not in wb.sheetnames:
        return
    ws = wb["Overview"]

    ws["A1"].value = f"BA Checklist — {project_name}"
    ws["A1"].font = Font(bold=True, size=14)

    # Rewrite projects table area (starting at A16)
    for r in range(16, 30):
        for c in range(1, 4):
            ws.cell(row=r, column=c).value = None

    ws["A15"].value = "Project"
    ws["A15"].font = Font(bold=True)

    ws["A16"].value = project_name
    ws["B16"].value = deliverables_path
    ws["C16"].value = "Project-specific checklist (standards-based)."


def rewrite_evidence_links(wb, project_name: str, deliverables_path: str) -> None:
    if "Evidence Links" not in wb.sheetnames:
        return
    ws = wb["Evidence Links"]

    # Clear existing rows after header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    ws.append([project_name, deliverables_path, "Paste links to PDFs/diagrams/SQL/UAT/RTM that support each checklist item."])


def fill_project_sheet(ws, project: str, mapping: dict[str, dict[str, tuple[str, str, str]]]) -> None:
    # Project columns in template
    # Foodora: F/G, Retail: H/I, SAP: J/K, Notes: L
    col_map = {
        "Foodora": {"status": "F", "evidence": "G"},
        "Retail": {"status": "H", "evidence": "I"},
        "SAP": {"status": "J", "evidence": "K"},
    }

    status_col = col_map[project]["status"]
    evidence_col = col_map[project]["evidence"]

    for row in range(2, ws.max_row + 1):
        item = ws[f"C{row}"].value
        if not item or item not in mapping:
            continue

        status, evidence, note = mapping[item][project]

        ws[f"{status_col}{row}"].value = status
        if evidence:
            set_cell_hyperlink(ws[f"{evidence_col}{row}"], evidence, "Evidence")
        else:
            ws[f"{evidence_col}{row}"].value = ""
            clear_hyperlink(ws[f"{evidence_col}{row}"])

        # Keep only this project's notes
        ws[f"L{row}"].value = note or ""
        ws[f"L{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Clear other project columns to avoid confusion
        for other in ["F", "G", "H", "I", "J", "K"]:
            if other in (status_col, evidence_col):
                continue
            ws[f"{other}{row}"].value = ""
            clear_hyperlink(ws[f"{other}{row}"])


def reduce_to_single_project_columns(ws, project: str) -> None:
    # Template columns:
    # A Standard, B Area, C Item, D Definition, E Priority,
    # F Foodora Status, G Foodora Evidence,
    # H Retail Status, I Retail Evidence,
    # J SAP Status, K SAP Evidence,
    # L Notes
    # We reduce to:
    # A Standard, B Area, C Item, D Definition, E Priority, F Status, G Evidence, H Notes

    if project == "Foodora":
        # Delete Retail + SAP columns (K,J,I,H)
        ws.delete_cols(11)  # K
        ws.delete_cols(10)  # J
        ws.delete_cols(9)   # I
        ws.delete_cols(8)   # H
    elif project == "Retail":
        # Delete SAP columns first (K,J), then Foodora (G,F)
        ws.delete_cols(11)  # K
        ws.delete_cols(10)  # J
        ws.delete_cols(7)   # G
        ws.delete_cols(6)   # F
    elif project == "SAP":
        # Delete Retail (I,H) and Foodora (G,F)
        ws.delete_cols(9)   # I
        ws.delete_cols(8)   # H
        ws.delete_cols(7)   # G
        ws.delete_cols(6)   # F

    # Rename headers for clarity
    ws["F1"].value = "Status"
    ws["G1"].value = "Evidence"
    ws["H1"].value = "Notes"


def build_mappings():
    # Mapping keys must match 'Checklist Item' values in the template.
    babok_map: dict[str, dict[str, tuple[str, str, str]]] = {
        "BA approach defined": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
        "Stakeholder analysis": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/Stakeholder Analysis Matrix.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Stakeholder_Engagement_Approach_and_RACI.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Stakeholders_RACI.pdf", ""),
        },
        "Governance & change control": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", "No explicit change-request log in portfolio"),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", "Change governance not explicitly documented"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", "Change governance not explicitly documented"),
        },
        "Elicitation plan": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", "Elicitation plan not separated as a standalone doc"),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Stakeholder_Engagement_Approach_and_RACI.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_BRD.pdf", ""),
        },
        "Elicitation results confirmed": {
            "Foodora": ("No", "", "Meeting minutes/confirmations not included"),
            "Retail": ("No", "", "Meeting minutes/confirmations not included"),
            "SAP": ("No", "", "Meeting minutes/confirmations not included"),
        },
        "Traceability": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Prioritization approach": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", "Priority method not formalized"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", "Priority method not formalized"),
        },
        "Requirement approvals": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Foodora_UAT_Signoff_Form.pdf", "Using UAT sign-off as approval evidence"),
            "Retail": ("No", "", "No explicit sign-off/approval artifact"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Summary_Report_IEEE829.pdf", "Using UAT summary as validation evidence"),
        },
        "Current state analysis (AS-IS)": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Business Case.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_BRD.pdf", ""),
        },
        "Future state (TO-BE) + success metrics": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "Options analysis": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Business Case.pdf", "Options comparison not expanded in a dedicated section"),
            "Retail": ("No", "", "Options analysis not included"),
            "SAP": ("No", "", "Options analysis not included"),
        },
        "Functional requirements + AC": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Use Case Description.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Process model": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/BPMN.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/diagrams/Retail_System_BPMN_L0.drawio.png", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/diagrams/SAP_ERP_BPMN_L0.drawio.png", ""),
        },
        "Data model": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/ERD - Foodora MVP.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/diagrams/Retail_System_Logical_ERD.drawio.png", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/diagrams/SAP_ERP_Logical_ERD.drawio.png", ""),
        },
        "Non-functional requirements (NFR)": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "UAT plan & scenarios": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/UAT_Test_Cases_Template.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_UAT_Test_Cases.xlsx", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Cases.xlsx", ""),
        },
        "Post-implementation measurement": {
            "Foodora": ("Partial", "projects/foodora-mvp/project-overview.html", "No real post-go-live metrics in portfolio"),
            "Retail": ("Partial", "projects/Retail_System/project-overview.html", "No real post-go-live metrics in portfolio"),
            "SAP": ("Partial", "projects/SAP_ERP/project-overview.html", "No real post-go-live metrics in portfolio"),
        },
    }

    iso_map: dict[str, dict[str, tuple[str, str, str]]] = {
        "Atomic": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Use Case Description.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Unambiguous": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Complete": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "Consistent": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Feasible": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/System Architecture Overview.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
        "Verifiable/Testable": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/UAT_Test_Cases_Template.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_UAT_Test_Cases.xlsx", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Cases.xlsx", ""),
        },
        "Traceable": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Prioritized": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
    }

    return babok_map, iso_map


def make_project_checklist(template_path: Path, out_path: Path, project: str, project_name: str, deliverables_path: str) -> None:
    babok_map, iso_map = build_mappings()

    wb = load_workbook(template_path)

    rewrite_overview(wb, project_name, deliverables_path)
    rewrite_evidence_links(wb, project_name, deliverables_path)

    for sheet_name, mapping in [("BABOK v3", babok_map), ("ISO 29148", iso_map)]:
        ws = wb[sheet_name]
        fill_project_sheet(ws, project, mapping)
        reduce_to_single_project_columns(ws, project)

        # Wrap text for Definition and Notes
        for row in range(2, ws.max_row + 1):
            ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f"H{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    template = Path("assets/docs/BA_Checklist_Template_BABOK_ISO.xlsx")

    targets = [
        (
            "Foodora",
            "Foodora MVP",
            "projects/foodora-mvp/deliverables.html",
            Path("assets/docs/BA_Checklist_Foodora_MVP.xlsx"),
        ),
        (
            "Retail",
            "Retail System",
            "projects/Retail_System/deliverables.html",
            Path("assets/docs/BA_Checklist_Retail_System.xlsx"),
        ),
        (
            "SAP",
            "SAP ERP (Mock)",
            "projects/SAP_ERP/deliverables.html",
            Path("assets/docs/BA_Checklist_SAP_ERP.xlsx"),
        ),
    ]

    for project, name, deliverables, out in targets:
        make_project_checklist(template, out, project, name, deliverables)


if __name__ == "__main__":
    main()
