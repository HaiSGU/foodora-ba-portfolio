from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


TODAY = date.today().isoformat()


BABOK_ITEMS = [
    # Planning & Monitoring
    {
        "standard": "BABOK v3",
        "area": "Business Analysis Planning & Monitoring",
        "item": "BA approach defined",
        "description": "Define BA approach (agile/waterfall/hybrid), deliverables, cadence, and working agreements.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Business Analysis Planning & Monitoring",
        "item": "Stakeholder analysis",
        "description": "Identify stakeholders, power/interest, roles; create engagement plan.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Business Analysis Planning & Monitoring",
        "item": "Governance & change control",
        "description": "Decision log, change requests, impact analysis, and approval/sign-off gates.",
        "priority": "Core",
    },
    # Elicitation
    {
        "standard": "BABOK v3",
        "area": "Elicitation & Collaboration",
        "item": "Elicitation plan",
        "description": "Plan workshops/interviews; define objectives, participants, and questions.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Elicitation & Collaboration",
        "item": "Elicitation results confirmed",
        "description": "Meeting notes / outcomes reviewed and confirmed by key stakeholders.",
        "priority": "Core",
    },
    # Requirements lifecycle
    {
        "standard": "BABOK v3",
        "area": "Requirements Life Cycle Management",
        "item": "Traceability",
        "description": "Trace objectives → requirements → design artifacts → test cases.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Requirements Life Cycle Management",
        "item": "Prioritization approach",
        "description": "Define prioritization method (MoSCoW/WSJF/value-risk) and rationale.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Requirements Life Cycle Management",
        "item": "Requirement approvals",
        "description": "Define who approves which requirement types and when (baseline control).",
        "priority": "Plus",
    },
    # Strategy analysis
    {
        "standard": "BABOK v3",
        "area": "Strategy Analysis",
        "item": "Current state analysis (AS-IS)",
        "description": "Pain points, root causes, constraints.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Strategy Analysis",
        "item": "Future state (TO-BE) + success metrics",
        "description": "Target outcomes and measurable KPIs; what good looks like.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Strategy Analysis",
        "item": "Options analysis",
        "description": "Compare solution options and recommend with rationale (value/risk/effort).",
        "priority": "Plus",
    },
    # Requirements analysis & design definition
    {
        "standard": "BABOK v3",
        "area": "Requirements Analysis & Design Definition",
        "item": "Functional requirements + AC",
        "description": "User stories/use cases with acceptance criteria (testable).",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Requirements Analysis & Design Definition",
        "item": "Process model",
        "description": "Flow/activity/BPMN for core scenarios; includes exceptions.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Requirements Analysis & Design Definition",
        "item": "Data model",
        "description": "Logical ERD + key entities/relationships + business rules.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Requirements Analysis & Design Definition",
        "item": "Non-functional requirements (NFR)",
        "description": "Performance, security, audit/logging, availability, usability.",
        "priority": "Core",
    },
    # Solution evaluation
    {
        "standard": "BABOK v3",
        "area": "Solution Evaluation",
        "item": "UAT plan & scenarios",
        "description": "UAT approach and scenarios for happy path + edge cases.",
        "priority": "Core",
    },
    {
        "standard": "BABOK v3",
        "area": "Solution Evaluation",
        "item": "Post-implementation measurement",
        "description": "Define how to measure outcomes after launch (KPIs/monitoring).",
        "priority": "Plus",
    },
]


ISO_29148_QUALITY = [
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Atomic",
        "description": "One requirement expresses one thing (no AND/OR bundles).",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Unambiguous",
        "description": "Only one interpretation; avoid vague words (fast, user-friendly, etc.).",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Complete",
        "description": "Includes conditions, exceptions, and necessary rules.",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Consistent",
        "description": "No conflicts with other requirements; terminology consistent.",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Feasible",
        "description": "Realistic within constraints (time, budget, tech, policy).",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Verifiable/Testable",
        "description": "Can be tested/inspected/verified with objective criteria.",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Traceable",
        "description": "Has a unique ID; links to source and downstream artifacts.",
        "priority": "Core",
    },
    {
        "standard": "ISO/IEC/IEEE 29148",
        "area": "Requirement quality",
        "item": "Prioritized",
        "description": "Priority recorded (MoSCoW/WSJF/value-risk), not all 'must'.",
        "priority": "Plus",
    },
]


PROJECTS = [
    {
        "key": "Foodora MVP",
        "deliverables": "projects/foodora-mvp/deliverables.html",
        "notes": "Strong: ERD + SQL pack + UAT template + traceability matrix.",
    },
    {
        "key": "Retail System",
        "deliverables": "projects/Retail_System/deliverables.html",
        "notes": "Strong: diagrams library + SQL pack + retail operational flows.",
    },
    {
        "key": "SAP ERP",
        "deliverables": "projects/SAP_ERP/deliverables.html",
        "notes": "Strong: O2C controls/monitoring + SQL pack + UAT cases.",
    },
]


HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def add_validation(ws, col_letter, start_row, end_row, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def main():
    wb = Workbook()

    # Overview
    ws0 = wb.active
    ws0.title = "Overview"
    ws0["A1"].value = "BA Checklist Template"
    ws0["A1"].font = TITLE_FONT
    ws0["A3"].value = "Standards"
    ws0["A3"].font = BOLD
    ws0["A4"].value = "- BABOK v3 (IIBA) knowledge areas"
    ws0["A5"].value = "- ISO/IEC/IEEE 29148 requirement quality criteria"

    ws0["A7"].value = "How to use"
    ws0["A7"].font = BOLD
    ws0["A8"].value = "1) For each checklist row: set Status per project (Yes/Partial/No/N/A)."
    ws0["A9"].value = "2) Paste evidence links (deliverables, PDFs, diagrams, SQL) in Evidence columns."
    ws0["A10"].value = "3) Use Notes to clarify scope/constraints (N/A is OK when no go-live)."

    ws0["A12"].value = "Generated"
    ws0["A12"].font = BOLD
    ws0["A13"].value = TODAY

    ws0["A15"].value = "Projects"
    ws0["A15"].font = BOLD
    r = 16
    for p in PROJECTS:
        ws0[f"A{r}"].value = p["key"]
        ws0[f"B{r}"].value = p["deliverables"]
        ws0[f"C{r}"].value = p["notes"]
        r += 1
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 45
    ws0.column_dimensions["C"].width = 70

    # BABOK sheet
    ws1 = wb.create_sheet("BABOK v3")
    headers = [
        "Standard",
        "Knowledge Area",
        "Checklist Item",
        "Definition",
        "Priority",
        "Foodora Status",
        "Foodora Evidence",
        "Retail Status",
        "Retail Evidence",
        "SAP Status",
        "SAP Evidence",
        "Notes",
    ]
    ws1.append(headers)
    style_header(ws1)

    for it in BABOK_ITEMS:
        ws1.append(
            [
                it["standard"],
                it["area"],
                it["item"],
                it["description"],
                it["priority"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    last_row = ws1.max_row
    add_validation(ws1, "F", 2, last_row, ["Yes", "Partial", "No", "N/A"])
    add_validation(ws1, "H", 2, last_row, ["Yes", "Partial", "No", "N/A"])
    add_validation(ws1, "J", 2, last_row, ["Yes", "Partial", "No", "N/A"])

    for col in ["D", "G", "I", "K", "L"]:
        for cell in ws1[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ISO sheet
    ws2 = wb.create_sheet("ISO 29148")
    ws2.append(headers)
    style_header(ws2)

    for it in ISO_29148_QUALITY:
        ws2.append(
            [
                it["standard"],
                it["area"],
                it["item"],
                it["description"],
                it["priority"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    last_row2 = ws2.max_row
    add_validation(ws2, "F", 2, last_row2, ["Yes", "Partial", "No", "N/A"])
    add_validation(ws2, "H", 2, last_row2, ["Yes", "Partial", "No", "N/A"])
    add_validation(ws2, "J", 2, last_row2, ["Yes", "Partial", "No", "N/A"])

    for col in ["D", "G", "I", "K", "L"]:
        for cell in ws2[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Evidence quick-links sheet
    ws3 = wb.create_sheet("Evidence Links")
    ws3.append(["Project", "Deliverables page", "Suggested evidence to paste"])
    style_header(ws3)
    for p in PROJECTS:
        ws3.append(
            [
                p["key"],
                p["deliverables"],
                "Paste links to diagrams, PDFs, SQL viewer, traceability/UAT files.",
            ]
        )
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 65

    # Light formatting
    for ws in [ws1, ws2]:
        ws.row_dimensions[1].height = 26
        autosize(ws)

    out_path = Path("assets/docs/BA_Checklist_Template_BABOK_ISO.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


if __name__ == "__main__":
    main()
