from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def set_cell_hyperlink(cell, url: str, label: str | None = None) -> None:
    if label is not None:
        cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def fill_sheet(ws, mapping: dict[str, dict[str, tuple[str, str, str]]]) -> None:
    """mapping[item] => {
        'Foodora': (status, evidence_path, notes),
        'Retail': (status, evidence_path, notes),
        'SAP': (status, evidence_path, notes),
    }
    """

    # Header row is 1.
    for row in range(2, ws.max_row + 1):
        item = ws[f"C{row}"].value
        if not item or item not in mapping:
            continue

        food_status, food_ev, food_notes = mapping[item]["Foodora"]
        retail_status, retail_ev, retail_notes = mapping[item]["Retail"]
        sap_status, sap_ev, sap_notes = mapping[item]["SAP"]

        ws[f"F{row}"].value = food_status
        ws[f"H{row}"].value = retail_status
        ws[f"J{row}"].value = sap_status

        if food_ev:
            set_cell_hyperlink(ws[f"G{row}"], food_ev, "Evidence")
        if retail_ev:
            set_cell_hyperlink(ws[f"I{row}"], retail_ev, "Evidence")
        if sap_ev:
            set_cell_hyperlink(ws[f"K{row}"], sap_ev, "Evidence")

        # Merge notes into Notes column (L)
        notes_parts = [n for n in [food_notes, retail_notes, sap_notes] if n]
        if notes_parts:
            ws[f"L{row}"].value = " | ".join(notes_parts)

        # Wrap long cells
        for col in ["D", "L"]:
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    template = Path("assets/docs/BA_Checklist_Template_BABOK_ISO.xlsx")
    out = Path("assets/docs/BA_Checklist_BABOK_ISO_Filled.xlsx")

    wb = load_workbook(template)

    babok_map: dict[str, dict[str, tuple[str, str, str]]] = {
        "BA approach defined": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
        "Stakeholder analysis": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/Stakeholder Analysis Matrix.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Stakeholder_Engagement_Approach_and_RACI.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Stakeholders_RACI.pdf", ""),
        },
        "Governance & change control": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", "No explicit change-request log in portfolio"),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", "Change governance not explicitly documented"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", "Change governance not explicitly documented"),
        },
        "Elicitation plan": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", "Elicitation plan not separated as a standalone doc"),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Stakeholder_Engagement_Approach_and_RACI.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_BRD.pdf", ""),
        },
        "Elicitation results confirmed": {
            "Foodora": ("No", "", "Meeting minutes/confirmations not included"),
            "Retail": ("No", "", "Meeting minutes/confirmations not included"),
            "SAP": ("No", "", "Meeting minutes/confirmations not included"),
        },
        "Traceability": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Prioritization approach": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", "Priority method not formalized"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", "Priority method not formalized"),
        },
        "Requirement approvals": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Foodora_UAT_Signoff_Form.pdf", "Using UAT sign-off as approval evidence"),
            "Retail": ("No", "", "No explicit sign-off/approval artifact"),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Summary_Report_IEEE829.pdf", "Using UAT summary as validation evidence"),
        },
        "Current state analysis (AS-IS)": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Business Case.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_BRD.pdf", ""),
        },
        "Future state (TO-BE) + success metrics": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "Options analysis": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Business Case.pdf", "Options comparison not expanded in a dedicated section"),
            "Retail": ("No", "", "Options analysis not included"),
            "SAP": ("No", "", "Options analysis not included"),
        },
        "Functional requirements + AC": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Use Case Description.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Process model": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/BPMN.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/diagrams/Retail_System_BPMN_L0.drawio.png", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/diagrams/SAP_ERP_BPMN_L0.drawio.png", ""),
        },
        "Data model": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/images/diagrams/ERD - Foodora MVP.drawio.png", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/diagrams/Retail_System_Logical_ERD.drawio.png", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/diagrams/SAP_ERP_Logical_ERD.drawio.png", ""),
        },
        "Non-functional requirements (NFR)": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "UAT plan & scenarios": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/UAT_Test_Cases_Template.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_UAT_Test_Cases.xlsx", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Cases.xlsx", ""),
        },
        "Post-implementation measurement": {
            "Foodora": ("Partial", "projects/foodora-mvp/project-overview.html", "No real post-go-live metrics in portfolio"),
            "Retail": ("Partial", "projects/Retail_System/project-overview.html", "No real post-go-live metrics in portfolio"),
            "SAP": ("Partial", "projects/SAP_ERP/project-overview.html", "No real post-go-live metrics in portfolio"),
        },
    }

    iso_map: dict[str, dict[str, tuple[str, str, str]]] = {
        "Atomic": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Use Case Description.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Unambiguous": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_SRS.pdf", ""),
        },
        "Complete": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/PRD.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Requirements_Scope_and_NFRs.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_Requirements_Scope_NFRs.pdf", ""),
        },
        "Consistent": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Feasible": {
            "Foodora": ("Partial", "projects/foodora-mvp/assets/docs/System Architecture Overview.pdf", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/PRD_Lite.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
        "Verifiable/Testable": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/UAT_Test_Cases_Template.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_UAT_Test_Cases.xlsx", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_UAT_Test_Cases.xlsx", ""),
        },
        "Traceable": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Foodora_Traceability_Matrix.xlsx", ""),
            "Retail": ("Yes", "projects/Retail_System/assets/docs/IIBA_Requirements_Traceability_Matrix.pdf", ""),
            "SAP": ("Yes", "projects/SAP_ERP/assets/docs/SAP_ERP_RTM.pdf", ""),
        },
        "Prioritized": {
            "Foodora": ("Yes", "projects/foodora-mvp/assets/docs/Jira Backlog and Technical Stories.docx", ""),
            "Retail": ("Partial", "projects/Retail_System/assets/docs/Sample_User_Stories.pdf", ""),
            "SAP": ("Partial", "projects/SAP_ERP/assets/docs/SAP_ERP_PRD.pdf", ""),
        },
    }

    fill_sheet(wb["BABOK v3"], babok_map)
    fill_sheet(wb["ISO 29148"], iso_map)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


if __name__ == "__main__":
    main()
